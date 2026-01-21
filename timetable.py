# timetable.py / app.py
# FULL Streamlit system (ERP + CBS + CBS2) with:
# ✅ Lecturer code mapping (Lec Codes.xlsx)
# ✅ Group Name generation (coded groups OR Group Mapping Excel)
# ✅ Option B: Force override intake label even if mapping exists (checkbox)
# ✅ Editable table (st.data_editor) and download edited CSV/Excel
# ✅ CBS output to template: EntrySheet (Teachers) + CBS2 (Teacher1)
#
# requirements.txt:
#   streamlit
#   pandas
#   openpyxl

import re
from io import BytesIO
from datetime import datetime, timedelta

import pandas as pd
import streamlit as st
import openpyxl

st.set_page_config(page_title="Timetable → ERP + CBS Outputs", layout="wide")
st.title("📅 Timetable CSV → ERP Upload + CBS Template Output (EntrySheet + CBS2)")

# -----------------------------
# ERP output columns (Week1 style) + Group Name added
# -----------------------------
ERP_OUTPUT_COLS = [
    "Activity Id",
    "Day",
    "Hour",
    "Students Sets",
    "Group Name",
    "Subject",
    "Teachers",
    "Teacher1",
    "Activity Tags",
    "Room",
    "Comments",
]

# -----------------------------
# CBS template columns (from your CBSTimeTableTemplate EntrySheet)
# -----------------------------
CBS_COLS = [
    "Cal Id",
    "Course",
    "Course Variant",
    "Section",
    "Room",
    "Faculty",
    "Day",
    "From Time Slot",
    "To Time Slot",
    "AcademyLocationID",
    "isAllFaculties",
]

# -----------------------------
# Helpers
# -----------------------------
def norm(s) -> str:
    if pd.isna(s) or s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()

def split_teacher_codes(raw_teachers: str):
    if pd.isna(raw_teachers) or raw_teachers is None:
        return []
    return [t.strip() for t in str(raw_teachers).split("+") if t.strip()]

def build_code_to_username_map(lec_df: pd.DataFrame) -> dict:
    lec_df = lec_df.copy()
    lec_df.columns = [c.strip() for c in lec_df.columns]

    required = {"Code", "User name"}
    missing = [c for c in required if c not in lec_df.columns]
    if missing:
        raise ValueError(f"Lec Codes.xlsx missing columns: {missing}. Required: {sorted(required)}")

    lec_df["Code"] = lec_df["Code"].astype(str).map(norm).str.upper()
    lec_df["User name"] = lec_df["User name"].astype(str).map(norm)
    lec_df = lec_df.dropna(subset=["Code", "User name"]).drop_duplicates(subset=["Code"], keep="first")

    return dict(zip(lec_df["Code"], lec_df["User name"]))

def map_teachers(raw_teachers: str, code_map: dict):
    """
    Returns: main, teacher1, unknown(list), extras(list beyond 2)
    Unknown codes kept as-is
    """
    codes = split_teacher_codes(raw_teachers)
    mapped, unknown = [], []
    for c in codes:
        key = c.upper()
        if key in code_map:
            mapped.append(code_map[key])
        else:
            unknown.append(c)
            mapped.append(c)

    main = mapped[0] if len(mapped) >= 1 else ""
    t1 = mapped[1] if len(mapped) >= 2 else ""
    extras = mapped[2:] if len(mapped) > 2 else []
    return main, t1, unknown, extras

# -----------------------------
# Group Name from coded Students Sets (e.g., L5 CS -G4, L6 IT -A1)
# -----------------------------
def extract_groups_from_codes(students_sets: str):
    """
    Supports:
      L5 SE -G10
      L6 IT -A1
      L4 CS -C3
    Returns dict like {'CS':[4,17], 'SE':[10]}
    """
    out = {}
    s = norm(students_sets)
    if not s:
        return out

    for part in s.split("+"):
        part = part.strip()
        m = re.search(r"\bL\d+\s+([A-Z&]{2,})\s*-\s*[A-Z]*?(\d+)\b", part, flags=re.IGNORECASE)
        if not m:
            continue
        prog = m.group(1).upper()
        num = int(m.group(2))
        out.setdefault(prog, []).append(num)

    for prog in out:
        out[prog] = sorted(set(out[prog]))
    return out

def build_group_name_from_codes(students_sets: str, intake_label: str) -> str:
    """
    ✅ Single program:
        L5 Jan 26 CS- 13,14,15,16
    ✅ Multi program (intake label only once):
        L5 Jan 26 SE- 10 / CS- 13,14,15,16
    """
    groups = extract_groups_from_codes(students_sets)
    if not groups:
        return ""

    def nums_str(prog):
        return ",".join(str(n) for n in groups[prog])

    order = []
    for p in ["SE", "CS"]:
        if p in groups:
            order.append(p)
    for p in sorted(groups.keys()):
        if p not in order:
            order.append(p)

    if len(order) == 1:
        p = order[0]
        return f"{intake_label} {p}- {nums_str(p)}"

    parts = [f"{intake_label} {order[0]}- {nums_str(order[0])}"]
    for p in order[1:]:
        parts.append(f"{p}- {nums_str(p)}")
    return " / ".join(parts)

# -----------------------------
# Group Mapping Excel (when Students Sets has no codes)
# -----------------------------
def load_group_map(xlsx_file) -> pd.DataFrame:
    """
    Required columns:
      Students Sets | Level | Intake | Program | Groups
    Example:
      AI & DS STAGE 1 [JAN] | L4 | Jan 26 | CS | 4,17,18,19,20
    """
    df = pd.read_excel(xlsx_file)
    df.columns = [c.strip() for c in df.columns]

    required = {"Students Sets", "Level", "Intake", "Program", "Groups"}
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Group mapping file missing columns: {missing}. Required: {sorted(required)}")

    df["Students Sets"] = df["Students Sets"].astype(str).map(norm)
    df["Level"] = df["Level"].astype(str).map(norm)
    df["Intake"] = df["Intake"].astype(str).map(norm)
    df["Program"] = df["Program"].astype(str).map(norm).str.upper()
    df["Groups"] = df["Groups"].astype(str).map(norm)
    return df

def group_name_from_map(students_sets: str, gm: pd.DataFrame) -> str:
    """
    Returns:
      L4 Jan 26 CS- 4,17,18,19,20
    or multi program:
      L4 Jan 26 SE- 10 / CS- 13,14,15,16
    """
    s = norm(students_sets)
    sub = gm[gm["Students Sets"] == s]
    if sub.empty:
        return ""

    level = sub.iloc[0]["Level"]
    intake = sub.iloc[0]["Intake"]

    order = []
    for p in ["SE", "CS"]:
        if (sub["Program"] == p).any():
            order.append(p)
    for p in sorted(sub["Program"].unique()):
        if p not in order:
            order.append(p)

    if len(order) == 1:
        p = order[0]
        g = sub[sub["Program"] == p].iloc[0]["Groups"]
        return f"{level} {intake} {p}- {g}"

    parts = []
    first = order[0]
    g0 = sub[sub["Program"] == first].iloc[0]["Groups"]
    parts.append(f"{level} {intake} {first}- {g0}")
    for p in order[1:]:
        g = sub[sub["Program"] == p].iloc[0]["Groups"]
        parts.append(f"{p}- {g}")
    return " / ".join(parts)

# -----------------------------
# ✅ Option B helper: override intake prefix even if mapping exists
# -----------------------------
def override_intake_prefix(section_text: str, intake_label: str) -> str:
    """
    Replace leading intake like:
      L5 Jan 25 CS- ...
      L5 Jan 25 SE- 10 / CS- 13,14
    with the current intake_label.
    """
    s = norm(section_text)
    if not s:
        return s

    return re.sub(r"^L\d+\s+[A-Za-z]+\s+\d+\s+", intake_label + " ", s)

# -----------------------------
# General helpers
# -----------------------------
def ensure_columns(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    out = df.copy()
    for c in cols:
        if c not in out.columns:
            out[c] = ""
    return out[cols]

def df_to_excel_bytes(df: pd.DataFrame, sheet_name="ERP") -> bytes:
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return bio.getvalue()

# -----------------------------
# CBS helpers
# -----------------------------
def infer_course_variant(activity_tags: str) -> str:
    s = norm(activity_tags).upper()
    if "LAB" in s:
        return "Lab"
    if "TUT" in s:
        return "Tutorial"
    if "LEC" in s:
        return "Lecture"
    return ""

def parse_time_hhmm(t: str) -> datetime:
    return datetime.strptime(t.strip(), "%H:%M")

def fmt_time_hhmmss(dt: datetime) -> str:
    return dt.strftime("%H:%M:%S")

def build_sessions_for_cbs(df_hourly: pd.DataFrame) -> pd.DataFrame:
    """
    Merge hourly rows into a session row:
      start = first hour
      end   = last hour + 1 hour
    """
    df = df_hourly.copy()
    df["Hour"] = df["Hour"].map(norm)

    session_keys = [
        "Activity Id", "Day", "Students Sets", "Group Name", "Subject",
        "Room", "Teachers", "Teacher1", "Activity Tags"
    ]
    for c in session_keys:
        if c not in df.columns:
            df[c] = ""

    sessions = []
    for _, g in df.groupby(session_keys, dropna=False):
        hours = sorted([parse_time_hhmm(x) for x in g["Hour"].dropna().unique() if x])
        if not hours:
            continue
        start = hours[0]
        end = hours[-1] + timedelta(hours=1)

        row = g.iloc[0].to_dict()
        row["From Time Slot"] = fmt_time_hhmmss(start)
        row["To Time Slot"] = fmt_time_hhmmss(end)
        sessions.append(row)

    return pd.DataFrame(sessions)

def write_cbs_template_with_cbs2(template_bytes: bytes, cbs_df: pd.DataFrame, cbs2_df: pd.DataFrame) -> bytes:
    """
    Writes:
      - EntrySheet : main CBS table
      - CBS2       : Teacher1 rows
    """
    wb = openpyxl.load_workbook(BytesIO(template_bytes))

    if "EntrySheet" not in wb.sheetnames:
        raise ValueError("CBS template does not contain a sheet named 'EntrySheet'.")
    ws = wb["EntrySheet"]

    # clear existing rows
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # EntrySheet data
    for i, rec in enumerate(cbs_df.to_dict("records"), start=2):
        for j, col in enumerate(CBS_COLS, start=1):
            ws.cell(row=i, column=j).value = rec.get(col, "")

    # CBS2 sheet
    if "CBS2" in wb.sheetnames:
        wb.remove(wb["CBS2"])
    ws2 = wb.create_sheet("CBS2")

    for j, col in enumerate(CBS_COLS, start=1):
        ws2.cell(row=1, column=j).value = col

    for i, rec in enumerate(cbs2_df.to_dict("records"), start=2):
        for j, col in enumerate(CBS_COLS, start=1):
            ws2.cell(row=i, column=j).value = rec.get(col, "")

    out_bio = BytesIO()
    wb.save(out_bio)
    return out_bio.getvalue()

# -----------------------------
# UI: uploads
# -----------------------------
u1, u2, u3, u4 = st.columns(4)
with u1:
    timetable_file = st.file_uploader("1) Timetable CSV (software export)", type=["csv"])
with u2:
    lec_codes_file = st.file_uploader("2) Lec Codes.xlsx", type=["xlsx"])
with u3:
    cbs_template_file = st.file_uploader("3) CBS Template Excel (optional)", type=["xlsx"])
with u4:
    group_map_file = st.file_uploader("4) Group Mapping Excel (optional)", type=["xlsx"])

st.subheader("Settings")
intake_label = st.text_input("Default intake label (used when coded groups exist)", value="L5 Jan 26")
force_intake_override = st.checkbox(
    "Force override intake label (even if Group Mapping Excel is uploaded)",
    value=True
)
auto_regen_group = st.checkbox("Auto-generate Group Name", value=True)
put_extras_in_comments = st.checkbox("If >2 lecturers, append extras into Comments", value=True)

st.subheader("CBS Settings")
default_is_all_faculties = st.selectbox("CBS: isAllFaculties", options=["FALSE", "TRUE"], index=0)

st.divider()

# -----------------------------
# Main processing
# -----------------------------
if timetable_file and lec_codes_file:
    # Load timetable
    try:
        df_raw = pd.read_csv(timetable_file)
    except Exception as e:
        st.error(f"Could not read timetable CSV: {e}")
        st.stop()

    # Load lecturer codes
    try:
        lec_df = pd.read_excel(lec_codes_file)
        code_map = build_code_to_username_map(lec_df)
    except Exception as e:
        st.error(f"Could not read Lec Codes.xlsx or build mapping: {e}")
        st.stop()

    # Load group map (optional)
    group_map_df = None
    if group_map_file:
        try:
            group_map_df = load_group_map(group_map_file)
        except Exception as e:
            st.error(f"Could not read Group Mapping Excel: {e}")
            st.stop()

    # Validate timetable columns
    required_cols = ["Activity Id", "Day", "Hour", "Students Sets", "Subject", "Teachers", "Activity Tags", "Room"]
    missing = [c for c in required_cols if c not in df_raw.columns]
    if missing:
        st.error(f"Timetable CSV missing columns: {missing}")
        st.stop()

    out = df_raw.copy()
    out["Students Sets"] = out["Students Sets"].map(norm)

    if "Comments" not in out.columns:
        out["Comments"] = ""

    # Lecturer mapping
    unknown_all = set()
    extra_total = 0
    teachers_main, teachers_1, comments = [], [], []

    for _, row in out.iterrows():
        main, t1, unknown, extras = map_teachers(row["Teachers"], code_map)
        teachers_main.append(main)
        teachers_1.append(t1)

        for u in unknown:
            unknown_all.add(u)

        comment = norm(row.get("Comments", ""))

        if extras and put_extras_in_comments:
            extra_total += len(extras)
            extra_txt = "Extra lecturers: " + " + ".join(extras)
            comment = (comment + " | " + extra_txt).strip(" |") if comment else extra_txt

        comments.append(comment)

    out["Teachers"] = teachers_main
    out["Teacher1"] = teachers_1
    out["Comments"] = comments

    # Group Name generation (with override option)
    def final_group_name(students_sets: str) -> str:
        s = norm(students_sets)

        # 1) mapping file
        if group_map_df is not None:
            mapped = group_name_from_map(s, group_map_df)
            if mapped:
                if force_intake_override:
                    return override_intake_prefix(mapped, intake_label)
                return mapped

        # 2) coded groups
        parsed = build_group_name_from_codes(s, intake_label=intake_label)
        if parsed:
            return parsed

        # 3) fallback
        return f"{intake_label} {s}".strip()

    if auto_regen_group:
        out["Group Name"] = out["Students Sets"].apply(final_group_name)
    else:
        if "Group Name" not in out.columns:
            out["Group Name"] = out["Students Sets"].map(norm)

    # ERP output
    out_erp = ensure_columns(out, ERP_OUTPUT_COLS)
    st.success("✅ ERP table generated. Edit below before downloading.")

    if unknown_all:
        st.warning("Lecturer codes not found (kept as-is): " + ", ".join(sorted(unknown_all)))

    # Editable preview
    st.subheader("Edit before download")
    edited_erp = st.data_editor(out_erp, use_container_width=True, num_rows="dynamic", hide_index=True)

    # Downloads (NO broken strings)
    st.subheader("Download edited ERP output")
    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            "⬇️ Download ERP CSV (edited)",
            data=edited_erp.to_csv(index=False).encode("utf-8"),
            file_name="ERP_ready_week_edited.csv",
            mime="text/csv",
        )
    with c2:
        st.download_button(
            "⬇️ Download ERP Excel (edited)",
            data=df_to_excel_bytes(edited_erp, sheet_name="ERP"),
            file_name="ERP_ready_week_edited.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # CBS output
    st.divider()
    st.subheader("CBS Template Output (EntrySheet + CBS2)")

    sessions = build_sessions_for_cbs(edited_erp)
    if sessions.empty:
        st.error("Could not build CBS sessions. Check Hour format (must be 08:30, 09:30...).")
        st.stop()

    cbs_main_rows = []
    cbs2_rows = []

    for _, r in sessions.iterrows():
        base = {
            "Cal Id": "",
            "Course": norm(r.get("Subject", "")),
            "Course Variant": infer_course_variant(r.get("Activity Tags", "")),
            "Section": norm(r.get("Group Name", "")),
            "Room": norm(r.get("Room", "")),
            "Faculty": norm(r.get("Teachers", "")),
            "Day": norm(r.get("Day", "")),
            "From Time Slot": r.get("From Time Slot", ""),
            "To Time Slot": r.get("To Time Slot", ""),
            "AcademyLocationID": "",
            "isAllFaculties": default_is_all_faculties,
        }
        cbs_main_rows.append(base)

        t1 = norm(r.get("Teacher1", ""))
        if t1:
            base2 = base.copy()
            base2["Faculty"] = t1
            cbs2_rows.append(base2)

    cbs_df = pd.DataFrame(cbs_main_rows)
    cbs2_df = pd.DataFrame(cbs2_rows)

    for c in CBS_COLS:
        if c not in cbs_df.columns:
            cbs_df[c] = ""
        if c not in cbs2_df.columns:
            cbs2_df[c] = ""
    cbs_df = cbs_df[CBS_COLS].copy()
    cbs2_df = cbs2_df[CBS_COLS].copy()

    st.write("✅ CBS Preview (EntrySheet):")
    st.dataframe(cbs_df, use_container_width=True)

    st.write("✅ CBS2 Preview (Teacher1 duplicates):")
    st.dataframe(cbs2_df, use_container_width=True)

    if cbs_template_file:
        try:
            filled_bytes = write_cbs_template_with_cbs2(
                cbs_template_file.getvalue(),
                cbs_df,
                cbs2_df,
            )
            st.download_button(
                "⬇️ Download CBS Filled Template (EntrySheet + CBS2)",
                data=filled_bytes,
                file_name="CBS_Timetable_Filled.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Could not fill CBS template: {e}")
    else:
        st.info("Upload CBS Template Excel to download filled template.")

else:
    st.info("Upload at least: Timetable CSV + Lec Codes.xlsx")
